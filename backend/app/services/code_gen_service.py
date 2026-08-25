import re
from app.schemas.plan import ActionPlanPayload


class CodeGenService:
    @staticmethod
    def generate_pandas_script(plan: ActionPlanPayload, input_file_path: str = "input.xlsx", output_file_path: str = "output.xlsx") -> str:
        """
        Synthesizes a clean, deterministic Python Pandas script based on the verified ActionPlanPayload.
        The generated script reads from input_file_path and saves to output_file_path.
        """
        # Check if plan contains purely text replacement / cell update operations
        is_pure_cell_update = len(plan.operations) > 0 and all(
            op.type in ["replace_text", "replace_value", "replace", "update", "update_value"]
            for op in plan.operations
        )

        if is_pure_cell_update:
            lines = [
                "import openpyxl",
                "import pandas as pd",
                "",
                "# 1. Load input paths",
                f"input_path = {repr(input_file_path)}",
                f"output_path = {repr(output_file_path)}",
                "",
                "if input_path.endswith('.csv'):",
                "    df = pd.read_csv(input_path)",
            ]
            for op in plan.operations:
                params = op.params or {}
                old_val = params.get("old_value") or params.get("from")
                new_val = params.get("new_value") or params.get("to")
                if old_val is not None and new_val is not None:
                    lines.append(f"    df = df.astype(str).replace({repr(str(old_val))}, {repr(str(new_val))})")
            lines.extend([
                "    df.to_csv(output_path, index=False)",
                "else:",
                "    # OpenPyXL In-Place Cell Mutation (Preserves 100% of formatting, merged headers & layout)",
                "    wb = openpyxl.load_workbook(input_path)",
                "    for ws in wb.worksheets:",
                "        for row in ws.iter_rows():",
                "            for cell in row:",
                "                if cell.value is not None:",
                "                    val_str = str(cell.value)",
            ])
            for op in plan.operations:
                params = op.params or {}
                old_val = params.get("old_value") or params.get("from")
                new_val = params.get("new_value") or params.get("to")
                if old_val is not None and new_val is not None:
                    lines.extend([
                        f"                    if {repr(str(old_val))} in val_str:",
                        f"                        cell.value = val_str.replace({repr(str(old_val))}, {repr(str(new_val))})"
                    ])
            lines.extend([
                "    wb.save(output_path)",
                "",
                "print('In-place cell update execution successfully completed.')"
            ])
            return "\n".join(lines)

        # Standard Pandas Transformation Pipeline for Structural Modifications
        lines = [
            "import pandas as pd",
            "",
            "# 1. Load input workbook",
            f"input_path = {repr(input_file_path)}",
            f"output_path = {repr(output_file_path)}",
            "",
            "# Helper function for smart header auto-detection",
            "def smart_read(file_path, s_name=None):",
            "    try:",
            "        if file_path.endswith('.csv'):",
            "            df_raw = pd.read_csv(file_path, header=None, nrows=10)",
            "        else:",
            "            df_raw = pd.read_excel(file_path, sheet_name=s_name, header=None, nrows=10)",
            "        best_row = 0",
            "        max_cnt = -1",
            "        for r in range(min(10, len(df_raw))):",
            "            cnt = len([v for v in df_raw.iloc[r].dropna() if str(v).strip() != ''])",
            "            if cnt > max_cnt:",
            "                max_cnt = cnt",
            "                best_row = r",
            "        if file_path.endswith('.csv'):",
            "            return pd.read_csv(file_path, header=best_row)",
            "        return pd.read_excel(file_path, sheet_name=s_name, header=best_row)",
            "    except Exception:",
            "        if file_path.endswith('.csv'):",
            "            return pd.read_csv(file_path)",
            "        return pd.read_excel(file_path, sheet_name=s_name)",
            "",
            "# Load sheets into dictionary using smart header detection",
            "if input_path.endswith('.csv'):",
            "    sheets_dict = {'Sheet1': smart_read(input_path)}",
            "else:",
            "    excel_file = pd.ExcelFile(input_path)",
            "    sheets_dict = {s: smart_read(input_path, s_name=s) for s in excel_file.sheet_names}",
            "",
            "# Primary Sheet Reference",
            "first_sheet_key = list(sheets_dict.keys())[0]",
            "df = sheets_dict[first_sheet_key]",
            "",
            "# 2. Apply Action Plan Operations",
        ]

        for idx, op in enumerate(plan.operations):
            lines.append(f"# Operation {idx + 1}: {op.description}")
            params = op.params or {}

            if op.type in ["replace_text", "replace_value", "replace", "update", "update_value"]:
                col = params.get("column")
                old_val = params.get("old_value") or params.get("from")
                new_val = params.get("new_value") or params.get("to")
                if old_val is not None and new_val is not None:
                    if col and col in ["df.columns"]:
                        lines.append(f"if {repr(col)} in df.columns:")
                        lines.append(f"    df[{repr(col)}] = df[{repr(col)}].astype(str).replace({repr(str(old_val))}, {repr(str(new_val))})")
                    else:
                        lines.append(f"df = df.replace({repr(str(old_val))}, {repr(str(new_val))})")
                        lines.append(f"df = df.replace({repr(old_val)}, {repr(new_val)})")


            elif op.type == "search_filter" or "keyword" in params:
                kw = params.get("keyword") or params.get("value")
                if kw:
                    lines.append(f"df = df[df.astype(str).apply(lambda r: r.str.contains({repr(str(kw))}, case=False, na=False)).any(axis=1)]")

            elif op.type == "filter":
                col = params.get("column")
                operator = params.get("operator", "==")
                value = params.get("value")
                if col:
                    if isinstance(value, list) or operator in ["isin", "in"]:
                        val_list = [str(v).strip().lower() for v in value] if isinstance(value, list) else [str(value).strip().lower()]
                        lines.append(f"df = df[df[{repr(col)}].astype(str).str.strip().str.lower().isin({repr(val_list)})]")
                    elif operator in [">", "<", ">=", "<="]:
                        try:
                            num_val = float(str(value).replace(',', ''))
                            lines.append(f"clean_series = df[{repr(col)}].astype(str).str.replace(r'[^0-9.-]', '', regex=True)")
                            lines.append(f"df = df[pd.to_numeric(clean_series, errors='coerce') {operator} {num_val}]")
                        except ValueError:
                            lines.append(f"df = df[df[{repr(col)}].astype(str).str.strip().str.lower() == {repr(str(value).strip().lower())}]")

                    elif isinstance(value, str):
                        if operator in ["==", "="]:
                            lines.append(f"df = df[df[{repr(col)}].astype(str).str.strip().str.lower() == {repr(str(value).strip().lower())}]")
                        else:
                            lines.append(f"df = df[df[{repr(col)}].astype(str).str.contains({repr(value)}, case=False, na=False)]")
                    else:
                        lines.append(f"df = df[df[{repr(col)}].astype(str) {operator} {repr(str(value))}]")


            elif op.type == "calculate_column":
                target = params.get("target_column") or "Calculated_Result"
                expr = params.get("expression")
                if expr:
                    lines.append("# Sanitize and coerce numeric columns before calculation execution")
                    lines.append("for col in df.columns:")
                    lines.append("    try:")
                    lines.append("        if any(k in str(col).lower() for k in ['id', 'code', 'date', 'name', 'status', 'category', 'method', 'department', 'state', 'section', 'result', 'stream', 'aging']): continue")
                    lines.append("        num_s = pd.to_numeric(df[col].astype(str).str.replace(r'[^0-9.-]', '', regex=True), errors='coerce')")
                    lines.append("        if num_s.notna().sum() > 0 and len(num_s.dropna()) >= len(df) * 0.5:")
                    lines.append("            df[col] = num_s")
                    lines.append("    except Exception:")
                    lines.append("        pass")

                    if ".mean()" in expr:
                        # Extract column inside df['col'].mean()
                        mean_m = re.search(r"df\[['\"](.+?)['\"]\]\.mean\(\)", expr)
                        if mean_m:
                            mcol = mean_m.group(1)
                            lines.append(f"if {repr(mcol)} in df.columns:")
                            lines.append(f"    df[{repr(mcol)}] = pd.to_numeric(df[{repr(mcol)}].astype(str).str.replace(r'[^0-9.-]', '', regex=True), errors='coerce')")
                    lines.append(f"df[{repr(target)}] = {expr}")




            elif op.type == "sort":
                sort_col = params.get("by") or params.get("column")
                ascending = params.get("ascending", True)
                if sort_col:
                    lines.append(f"# Sort table by {sort_col} ({'ascending' if ascending else 'descending'})")
                    lines.append(f"if {repr(sort_col)} in df.columns:")
                    lines.append(f"    # Ensure numeric columns are properly sorted numerically")
                    lines.append(f"    sort_series = df[{repr(sort_col)}].astype(str).str.replace(r'[^0-9.-]', '', regex=True)")
                    lines.append(f"    num_series = pd.to_numeric(sort_series, errors='coerce')")
                    lines.append(f"    if num_series.notna().sum() > 0 and len(num_series.dropna()) == len(df):")
                    lines.append(f"        df['_sort_key'] = num_series")
                    lines.append(f"        df = df.sort_values(by='_sort_key', ascending={ascending}).drop(columns=['_sort_key'])")
                    lines.append(f"    else:")
                    lines.append(f"        df = df.sort_values(by={repr(sort_col)}, ascending={ascending})")

            elif op.type == "create_sheet":

                sheet_name = params.get("sheet_name", "Summary_Sheet")
                lines.append(f"sheets_dict[{repr(sheet_name)}] = df.copy()")

            elif op.type == "rename_column":
                old_name = params.get("column") or params.get("old_name")
                new_name = params.get("target_column") or params.get("new_name")
                if old_name and new_name:
                    lines.append(f"df = df.rename(columns={{{repr(old_name)}: {repr(new_name)}}})")

            elif op.type == "delete_column":
                col = params.get("column")
                if col:
                    lines.append(f"df = df.drop(columns=[{repr(col)}], errors='ignore')")

            lines.append("")

        # Update primary sheet in dictionary & clean temporary processing columns
        lines.extend([
            "# Drop temporary processing columns before output export",
            "temp_cols = [c for c in df.columns if str(c).startswith('temp_')]",
            "if temp_cols:",
            "    df = df.drop(columns=temp_cols, errors='ignore')",
            "",
            "sheets_dict[first_sheet_key] = df",
            "",
            "# 3. Save transformed workbook",
            "if output_path.endswith('.csv'):",
            "    df.to_csv(output_path, index=False)",

            "else:",
            "    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:",
            "        for s_name, s_df in sheets_dict.items():",
            "            s_df.to_excel(writer, sheet_name=s_name, index=False)",
            "",
            "print('Transformation execution successfully completed.')"
        ])

        return "\n".join(lines)
